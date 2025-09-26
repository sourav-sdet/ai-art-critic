from flask import Flask, render_template, request
import os
from openpyxl import load_workbook, Workbook
from flask import Flask, render_template, request, flash
import uuid

app = Flask(__name__)
app.secret_key = 'supersecretkey'  # required for flash messages

EXCEL_FILE = "users.xlsx"

# Ensure Excel file exists and is valid
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["username", "userId"])
    wb.save(EXCEL_FILE)
else:
    try:
        wb = load_workbook(EXCEL_FILE)  # test if it's valid
    except Exception as e:
        print("Invalid Excel file detected, creating new one...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["username", "userId"])
        wb.save(EXCEL_FILE)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        username = request.form.get("username").strip()

        if not username:
            flash("Username cannot be empty!", "error")
            return render_template("index.html")

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Check if username exists
        existing_users = [cell.value for cell in ws["A"][1:]]  # skip header
        if username in existing_users:
            flash(f"Username '{username}' already exists!", "error")
        else:
            # Generate a Unique User Id
            userId = str(uuid.uuid4())
            ws.append([username,userId])
            wb.save(EXCEL_FILE)
            flash(f"Username '{username}' added successfully!", "success")

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
